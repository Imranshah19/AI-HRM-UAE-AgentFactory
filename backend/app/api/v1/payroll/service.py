"""
AI-HRMS — Payroll service layer.
All DB operations are tenant-scoped; every function receives tenant_id explicitly.
"""

from __future__ import annotations

import calendar
import io
import logging
import os
from datetime  import date, datetime
from typing    import Optional

from fastapi             import HTTPException, status
from sqlalchemy          import select, func, and_, update
from sqlalchemy.orm      import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Employee,
    Department,
    Designation,
    SalaryStructure,
    BankDetails,
    AttendanceRecord,
    LeaveRequest,
    User,
)
from app.models.payroll import PayrollRun, PayrollRecord, TaxSlab

from app.api.v1.payroll.schemas import (
    PayrollFilterParams,
    PayrollRunCreate,
    PayrollRunResponse,
    PayrollRunDetailResponse,
    PayrollRecordResponse,
    EmployeeMinimal,
    PayslipData,
    AllowanceItem,
    DeductionItem,
    SalaryCalculationPreview,
    TaxSlabCreate,
    TaxSlabUpdate,
    TaxSlabResponse,
    PayrollApprovalRequest,
    BankFileEntry,
)
from app.api.v1.payroll.calculator import (
    TaxSlabData,
    SalaryStructureData,
    AttendanceSummaryData,
    LeaveSummaryData,
    calculate_employee_payroll,
    calculate_income_tax_breakdown,
    get_working_days,
)

logger = logging.getLogger(__name__)

PAYSLIP_DIR = os.environ.get("PAYSLIP_DIR", "/tmp/payslips")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _month_date_range(month: int, year: int) -> tuple[date, date]:
    _, last = calendar.monthrange(year, month)
    return date(year, month, 1), date(year, month, last)


def _str(v) -> str:
    return str(v) if v is not None else None


def _employee_minimal(emp: Employee) -> EmployeeMinimal:
    return EmployeeMinimal(
        id               = _str(emp.id),
        employee_code    = emp.employee_code,
        full_name        = f"{emp.first_name} {emp.last_name}",
        department_name  = emp.department.name if emp.department else None,
        designation_title= emp.designation.title if emp.designation else None,
        photo_url        = getattr(emp, "photo_url", None),
    )


def _record_response(record: PayrollRecord, emp: Employee) -> PayrollRecordResponse:
    d = {c.key: getattr(record, c.key) for c in record.__table__.columns}
    d["id"]             = _str(record.id)
    d["payroll_run_id"] = _str(record.payroll_run_id)
    d["employee_id"]    = _str(record.employee_id)
    d["employee"]       = _employee_minimal(emp)
    # Numeric fields stored as Decimal in DB — cast
    d["overtime_hours"]    = float(record.overtime_hours) if record.overtime_hours else None
    d["paid_leave_days"]   = float(record.paid_leave_days)
    d["unpaid_leave_days"] = float(record.unpaid_leave_days)
    return PayrollRecordResponse(**d)


# ─── Tax Slabs ────────────────────────────────────────────────────────────────

async def get_tax_slabs(tenant_id: str, year: int, db: AsyncSession) -> list[TaxSlab]:
    rows = await db.execute(
        select(TaxSlab)
        .where(TaxSlab.tenant_id == tenant_id, TaxSlab.year == year, TaxSlab.is_active == True)
        .order_by(TaxSlab.min_income)
    )
    return rows.scalars().all()


async def create_tax_slab(tenant_id: str, data: TaxSlabCreate, db: AsyncSession) -> TaxSlab:
    slab = TaxSlab(tenant_id=tenant_id, **data.model_dump())
    db.add(slab)
    await db.flush()
    await db.refresh(slab)
    await db.commit()
    return slab


async def update_tax_slab(
    tenant_id: str, slab_id: str, data: TaxSlabUpdate, db: AsyncSession
) -> TaxSlab:
    row = await db.execute(
        select(TaxSlab).where(TaxSlab.id == slab_id, TaxSlab.tenant_id == tenant_id)
    )
    slab = row.scalar_one_or_none()
    if not slab:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Tax slab not found.")
    for field_name, value in data.model_dump(exclude_none=True).items():
        setattr(slab, field_name, value)
    await db.commit()
    await db.refresh(slab)
    return slab


# ─── Salary Preview ───────────────────────────────────────────────────────────

async def preview_salary(
    tenant_id: str, employee_id: str, month: int, year: int, db: AsyncSession
) -> SalaryCalculationPreview:
    """Calculate salary for a preview without creating any payroll records."""
    emp, salary_struct, att_summary, leave_summary, tax_slab_list = await _fetch_payroll_inputs(
        tenant_id, employee_id, month, year, db
    )

    result = calculate_employee_payroll(
        employee_id = str(emp.id),
        salary      = salary_struct,
        attendance  = att_summary,
        leaves      = leave_summary,
        tax_slabs   = tax_slab_list,
        month       = month,
        year        = year,
    )

    annual_gross = result.gross_salary * 12
    tax_breakdown = calculate_income_tax_breakdown(annual_gross, tax_slab_list)

    return SalaryCalculationPreview(
        employee_id     = str(emp.id),
        employee_name   = f"{emp.first_name} {emp.last_name}",
        month           = month,
        year            = year,
        basic_salary    = result.basic_salary,
        total_allowances= result.total_allowances,
        gross_salary    = result.gross_salary,
        eobi_employee   = result.eobi_employee,
        eobi_employer   = result.eobi_employer,
        income_tax      = result.income_tax,
        loan_deduction  = result.loan_deduction,
        advance_deduction= result.advance_deduction,
        other_deductions_total= 0,
        total_deductions= result.total_deductions,
        net_salary      = result.net_salary,
        working_days    = result.working_days,
        present_days    = result.present_days,
        absent_days     = result.absent_days,
        overtime_hours  = result.overtime_hours,
        overtime_amount = result.overtime_amount,
        effective_tax_rate= tax_breakdown["effective_rate"],
        annual_gross    = annual_gross,
        annual_tax      = tax_breakdown["annual_tax"],
    )


# ─── Internal helpers for fetching calculation inputs ─────────────────────────

async def _fetch_payroll_inputs(
    tenant_id: str,
    employee_id: str,
    month: int,
    year: int,
    db: AsyncSession,
) -> tuple:
    """Fetch employee, salary, attendance, leave, tax slabs for one employee."""
    # Employee
    emp_row = await db.execute(
        select(Employee)
        .options(
            selectinload(Employee.department),
            selectinload(Employee.designation),
        )
        .where(Employee.id == employee_id, Employee.tenant_id == tenant_id)
    )
    emp = emp_row.scalar_one_or_none()
    if not emp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, f"Employee {employee_id} not found.")

    # Active salary structure (effective_to IS NULL)
    sal_row = await db.execute(
        select(SalaryStructure)
        .where(
            SalaryStructure.employee_id == employee_id,
            SalaryStructure.tenant_id   == tenant_id,
            SalaryStructure.effective_to == None,
        )
    )
    sal = sal_row.scalar_one_or_none()
    if not sal:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            f"No active salary structure for employee {employee_id}.",
        )

    # Attendance summary
    start_date, end_date = _month_date_range(month, year)
    att_rows = await db.execute(
        select(AttendanceRecord)
        .where(
            AttendanceRecord.employee_id == employee_id,
            AttendanceRecord.tenant_id   == tenant_id,
            AttendanceRecord.date.between(start_date, end_date),
        )
    )
    att_records = att_rows.scalars().all()

    working_days_in_month = get_working_days(month, year)
    present_days  = sum(1 for a in att_records if a.status in ("present", "late"))
    absent_days   = sum(1 for a in att_records if a.status == "absent")
    late_days     = sum(1 for a in att_records if a.status == "late")
    overtime_hrs  = float(sum(
        (a.overtime_hours or 0) for a in att_records if hasattr(a, "overtime_hours")
    ))

    att_summary = AttendanceSummaryData(
        working_days  = working_days_in_month,
        present_days  = present_days,
        absent_days   = absent_days,
        late_days     = late_days,
        overtime_hours= overtime_hrs,
        shift_hours   = 8.0,
    )

    # Leave summary
    leave_rows = await db.execute(
        select(LeaveRequest)
        .where(
            LeaveRequest.employee_id == employee_id,
            LeaveRequest.tenant_id   == tenant_id,
            LeaveRequest.status      == "approved",
            LeaveRequest.start_date  <= end_date,
            LeaveRequest.end_date    >= start_date,
        )
    )
    leave_records = leave_rows.scalars().all()
    paid_days   = 0.0
    unpaid_days = 0.0
    for lr in leave_records:
        # Use days field if available, else compute
        days = getattr(lr, "days", 1)
        if getattr(lr, "leave_type", None) and hasattr(lr.leave_type, "is_paid"):
            if lr.leave_type.is_paid:
                paid_days   += days
            else:
                unpaid_days += days
        else:
            paid_days += days  # default to paid

    leave_summary = LeaveSummaryData(
        paid_leave_days   = paid_days,
        unpaid_leave_days = unpaid_days,
    )

    salary_struct = SalaryStructureData(
        basic_salary           = sal.basic_salary,
        house_rent_allowance   = sal.house_rent_allowance or 0,
        medical_allowance      = sal.medical_allowance or 0,
        transport_allowance    = sal.transport_allowance or 0,
        fuel_allowance         = sal.fuel_allowance or 0,
        utility_allowance      = getattr(sal, "utility_allowance", 0) or 0,
        other_allowances       = sal.other_allowances or {},
        eobi_applicable        = sal.eobi_applicable,
        sessi_applicable       = getattr(sal, "sessi_applicable", False),
        income_tax_applicable  = sal.income_tax_applicable,
        loan_deduction         = sal.loan_deduction or 0,
        advance_deduction      = sal.advance_deduction or 0,
    )

    # Tax slabs
    tax_slabs_db = await get_tax_slabs(tenant_id, year, db)
    tax_slab_list = [
        TaxSlabData(
            min_income  = s.min_income,
            max_income  = s.max_income,
            fixed_tax   = s.fixed_tax,
            tax_rate    = float(s.tax_rate),
        )
        for s in tax_slabs_db
    ] if tax_slabs_db else None

    return emp, salary_struct, att_summary, leave_summary, tax_slab_list


# ─── Payroll Runs ─────────────────────────────────────────────────────────────

async def create_payroll_run(
    tenant_id: str, data: PayrollRunCreate, created_by: str, db: AsyncSession
) -> PayrollRun:
    """Create a PayrollRun and trigger async Celery processing."""
    # Guard: only one run per month/year/tenant
    existing = await db.execute(
        select(PayrollRun).where(
            PayrollRun.tenant_id == tenant_id,
            PayrollRun.month     == data.month,
            PayrollRun.year      == data.year,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Payroll for {data.month}/{data.year} has already been run for this organisation."
        )

    month_name = date(data.year, data.month, 1).strftime("%B")
    run = PayrollRun(
        tenant_id    = tenant_id,
        month        = data.month,
        year         = data.year,
        label        = f"{month_name} {data.year} Payroll",
        status       = "processing",
        processed_by = created_by,
        run_at       = datetime.utcnow(),
        notes        = data.notes,
    )
    db.add(run)
    await db.flush()
    await db.refresh(run)
    await db.commit()

    # Trigger Celery task (import here to avoid circular imports)
    try:
        from app.tasks.payroll_tasks import process_payroll_run
        task = process_payroll_run.delay(
            str(run.id),
            tenant_id,
            data.department_ids or [],
        )
        # Store task ID for polling
        await db.execute(
            update(PayrollRun)
            .where(PayrollRun.id == run.id)
            .values(celery_task_id=task.id)
        )
        await db.commit()
    except Exception as exc:
        logger.warning("Celery unavailable, payroll will process synchronously: %s", exc)

    await db.refresh(run)
    return run


async def get_payroll_runs(
    tenant_id: str, filters: PayrollFilterParams, db: AsyncSession
) -> dict:
    q = select(PayrollRun).where(PayrollRun.tenant_id == tenant_id)

    if filters.month:
        q = q.where(PayrollRun.month == filters.month)
    if filters.year:
        q = q.where(PayrollRun.year == filters.year)
    if filters.status:
        q = q.where(PayrollRun.status == filters.status)

    count_q  = select(func.count()).select_from(q.subquery())
    total    = (await db.execute(count_q)).scalar_one()

    offset   = (filters.page - 1) * filters.page_size
    q        = q.order_by(PayrollRun.year.desc(), PayrollRun.month.desc())
    q        = q.offset(offset).limit(filters.page_size)
    rows     = await db.execute(q)
    runs     = rows.scalars().all()

    return {"count": total, "results": runs}


async def get_payroll_run_by_id(
    tenant_id: str, run_id: str, db: AsyncSession
) -> PayrollRun:
    row = await db.execute(
        select(PayrollRun)
        .options(
            selectinload(PayrollRun.records).selectinload(PayrollRecord.employee).options(
                selectinload(Employee.department),
                selectinload(Employee.designation),
            )
        )
        .where(PayrollRun.id == run_id, PayrollRun.tenant_id == tenant_id)
    )
    run = row.scalar_one_or_none()
    if not run:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Payroll run not found.")
    return run


# ─── Approval / Rejection ─────────────────────────────────────────────────────

async def approve_payroll_run(
    tenant_id: str, run_id: str, approved_by: str, notes: Optional[str], db: AsyncSession
) -> PayrollRun:
    run = await _get_run(tenant_id, run_id, db)
    if run.status != "processed":
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            f"Only processed payroll runs can be approved. Current status: {run.status}",
        )
    run.status      = "approved"
    run.approved_by = approved_by
    run.approved_at = datetime.utcnow()
    if notes:
        run.notes = notes
    await db.commit()
    await db.refresh(run)

    # Generate payslips async
    try:
        from app.tasks.payroll_tasks import generate_all_payslips
        generate_all_payslips.delay(run_id, tenant_id)
    except Exception as exc:
        logger.warning("Could not queue payslip generation: %s", exc)

    return run


async def reject_payroll_run(
    tenant_id: str, run_id: str, notes: str, db: AsyncSession
) -> PayrollRun:
    if not notes or not notes.strip():
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Rejection reason is required.")

    run = await _get_run(tenant_id, run_id, db)
    if run.status not in ("processing", "processed"):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            f"Cannot reject a payroll run with status: {run.status}",
        )

    # Delete all records
    from sqlalchemy import delete as sql_delete
    await db.execute(
        sql_delete(PayrollRecord).where(PayrollRecord.payroll_run_id == run.id)
    )

    run.status = "cancelled"
    run.notes  = notes
    await db.commit()
    await db.refresh(run)
    return run


async def _get_run(tenant_id: str, run_id: str, db: AsyncSession) -> PayrollRun:
    row = await db.execute(
        select(PayrollRun).where(PayrollRun.id == run_id, PayrollRun.tenant_id == tenant_id)
    )
    run = row.scalar_one_or_none()
    if not run:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Payroll run not found.")
    return run


# ─── Payslips ─────────────────────────────────────────────────────────────────

async def get_employee_payslip_history(
    tenant_id: str, employee_id: str, db: AsyncSession
) -> list[PayrollRecord]:
    rows = await db.execute(
        select(PayrollRecord)
        .join(PayrollRun, PayrollRun.id == PayrollRecord.payroll_run_id)
        .where(
            PayrollRecord.employee_id == employee_id,
            PayrollRun.tenant_id      == tenant_id,
        )
        .order_by(PayrollRun.year.desc(), PayrollRun.month.desc())
    )
    return rows.scalars().all()


async def get_payslip_record(
    tenant_id: str, record_id: str, requesting_employee_id: Optional[str], db: AsyncSession
) -> PayrollRecord:
    row = await db.execute(
        select(PayrollRecord)
        .join(PayrollRun, PayrollRun.id == PayrollRecord.payroll_run_id)
        .where(
            PayrollRecord.id       == record_id,
            PayrollRun.tenant_id   == tenant_id,
        )
    )
    record = row.scalar_one_or_none()
    if not record:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Payslip not found.")

    # Employees can only access their own payslips
    if requesting_employee_id and str(record.employee_id) != str(requesting_employee_id):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "You can only access your own payslips.")

    return record


async def generate_payslip_pdf(record_id: str, db: AsyncSession) -> str:
    """
    Generate a PDF payslip for a PayrollRecord using ReportLab.
    Returns the file path.
    """
    row = await db.execute(
        select(PayrollRecord)
        .options(
            selectinload(PayrollRecord.employee).options(
                selectinload(Employee.department),
                selectinload(Employee.designation),
                selectinload(Employee.bank_details),
            ),
            selectinload(PayrollRecord.payroll_run),
        )
        .where(PayrollRecord.id == record_id)
    )
    record = row.scalar_one_or_none()
    if not record:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Payroll record not found.")

    emp = record.employee
    run = record.payroll_run

    os.makedirs(PAYSLIP_DIR, exist_ok=True)
    filename  = f"payslip_{emp.employee_code}_{run.year}_{run.month:02d}.pdf"
    file_path = os.path.join(PAYSLIP_DIR, filename)

    _generate_pdf(record, emp, run, file_path)

    # Update URL
    record.payslip_url = f"/payslips/{filename}"
    record.status      = "processed"
    await db.commit()

    return record.payslip_url


def _generate_pdf(record: PayrollRecord, emp: Employee, run: PayrollRun, file_path: str):
    """
    Generate payslip PDF with ReportLab.
    Falls back gracefully if reportlab is not installed.
    """
    try:
        from reportlab.lib.pagesizes  import A4
        from reportlab.lib.units      import cm
        from reportlab.lib            import colors
        from reportlab.lib.styles     import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus       import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
        from reportlab.lib.enums      import TA_CENTER, TA_RIGHT, TA_LEFT

        doc    = SimpleDocTemplate(file_path, pagesize=A4,
                                   topMargin=1.5*cm, bottomMargin=1.5*cm,
                                   leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = []

        # Header
        header_style = ParagraphStyle("header", fontSize=18, fontName="Helvetica-Bold",
                                      alignment=TA_CENTER, textColor=colors.HexColor("#1e293b"))
        sub_style    = ParagraphStyle("sub",    fontSize=10, fontName="Helvetica",
                                      alignment=TA_CENTER, textColor=colors.HexColor("#64748b"))
        normal       = ParagraphStyle("normal", fontSize=9,  fontName="Helvetica",
                                      textColor=colors.HexColor("#334155"))
        bold         = ParagraphStyle("bold",   fontSize=9,  fontName="Helvetica-Bold",
                                      textColor=colors.HexColor("#0f172a"))
        big_net      = ParagraphStyle("bignet", fontSize=22, fontName="Helvetica-Bold",
                                      alignment=TA_CENTER, textColor=colors.HexColor("#16a34a"))

        month_name  = date(run.year, run.month, 1).strftime("%B %Y")
        story.append(Paragraph("AI-HRMS", header_style))
        story.append(Paragraph(f"Salary Slip — {month_name}", sub_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e2e8f0")))
        story.append(Spacer(1, 0.3*cm))

        # Employee details table
        bank = next((b for b in (emp.bank_details or []) if b.is_primary), None)
        emp_data = [
            ["Employee Name",  f"{emp.first_name} {emp.last_name}",
             "Employee Code",  emp.employee_code],
            ["Department",     emp.department.name if emp.department else "—",
             "Designation",    emp.designation.title if emp.designation else "—"],
            ["Bank",           bank.bank_name if bank else "—",
             "Account No.",    bank.account_number if bank else "—"],
            ["IBAN",           bank.iban if bank else "—",
             "Pay Period",     month_name],
        ]
        emp_table = Table(emp_data, colWidths=[3.5*cm, 6.5*cm, 3.5*cm, 4.5*cm])
        emp_table.setStyle(TableStyle([
            ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME",    (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8.5),
            ("TEXTCOLOR",   (0, 0), (0, -1), colors.HexColor("#475569")),
            ("TEXTCOLOR",   (2, 0), (2, -1), colors.HexColor("#475569")),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("PADDING",     (0, 0), (-1, -1), 5),
        ]))
        story.append(emp_table)
        story.append(Spacer(1, 0.4*cm))

        # Earnings & deductions side by side
        earn_data  = [["Earnings", "Amount (PKR)"]]
        earn_data += [["Basic Salary", f"{record.basic_salary:,}"]]
        if record.house_rent_allowance:
            earn_data.append(["House Rent Allowance", f"{record.house_rent_allowance:,}"])
        if record.medical_allowance:
            earn_data.append(["Medical Allowance", f"{record.medical_allowance:,}"])
        if record.transport_allowance:
            earn_data.append(["Transport Allowance", f"{record.transport_allowance:,}"])
        if record.fuel_allowance:
            earn_data.append(["Fuel Allowance", f"{record.fuel_allowance:,}"])
        if record.other_allowances:
            for k, v in record.other_allowances.items():
                earn_data.append([k.replace("_", " ").title(), f"{v:,}"])
        earn_data.append(["Total Gross", f"{record.gross_salary:,}"])

        ded_data  = [["Deductions", "Amount (PKR)"]]
        if record.eobi_employee:
            ded_data.append(["EOBI (Employee 1%)", f"{record.eobi_employee:,}"])
        if record.income_tax:
            ded_data.append(["Income Tax", f"{record.income_tax:,}"])
        if record.loan_deduction:
            ded_data.append(["Loan Deduction", f"{record.loan_deduction:,}"])
        if record.advance_deduction:
            ded_data.append(["Advance Recovery", f"{record.advance_deduction:,}"])
        ded_data.append(["Total Deductions", f"{record.total_deductions:,}"])

        tbl_style = TableStyle([
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 8.5),
            ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1e293b")),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
            ("BACKGROUND",  (0, -1), (-1, -1), colors.HexColor("#f1f5f9")),
            ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
            ("ALIGN",       (1, 0), (1, -1),  "RIGHT"),
            ("PADDING",     (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
        ])

        earn_t = Table(earn_data, colWidths=[8*cm, 4*cm])
        earn_t.setStyle(tbl_style)
        ded_t  = Table(ded_data,  colWidths=[8*cm, 4*cm])
        ded_t.setStyle(tbl_style)

        two_col = Table([[earn_t, ded_t]], colWidths=[12.5*cm, 12.5*cm])
        two_col.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                                      ("LEFTPADDING",  (1, 0), (1, 0), 20)]))
        story.append(two_col)
        story.append(Spacer(1, 0.5*cm))

        # Net Salary (prominent)
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e2e8f0")))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(f"Net Salary: PKR {record.net_salary:,}", big_net))
        story.append(Spacer(1, 0.2*cm))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e2e8f0")))
        story.append(Spacer(1, 0.5*cm))

        # Attendance summary
        att_data = [
            ["Working Days", "Present Days", "Absent Days", "Paid Leave", "Overtime Hrs"],
            [
                str(record.working_days),
                str(record.present_days),
                str(record.absent_days),
                str(record.paid_leave_days),
                str(record.overtime_hours or 0),
            ],
        ]
        att_t = Table(att_data, colWidths=[3.6*cm]*5)
        att_t.setStyle(TableStyle([
            ("FONTNAME",   (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8.5),
            ("BACKGROUND", (0, 0), (-1, 0),  colors.HexColor("#334155")),
            ("TEXTCOLOR",  (0, 0), (-1, 0),  colors.white),
            ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("PADDING",    (0, 0), (-1, -1), 5),
        ]))
        story.append(att_t)
        story.append(Spacer(1, 1*cm))

        # Footer
        gen_date = datetime.utcnow().strftime("%d %b %Y, %H:%M UTC")
        footer   = ParagraphStyle("footer", fontSize=7.5, fontName="Helvetica",
                                  alignment=TA_CENTER, textColor=colors.HexColor("#94a3b8"))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e2e8f0")))
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            f"Generated on {gen_date} by AI-HRMS  |  This is a computer-generated document.",
            footer,
        ))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("________________________", footer))
        story.append(Paragraph("HR Department Signature", footer))

        doc.build(story)
        logger.info("Payslip PDF generated: %s", file_path)

    except ImportError:
        logger.warning("reportlab not installed — writing placeholder payslip text file.")
        with open(file_path.replace(".pdf", ".txt"), "w") as f:
            f.write(f"Payslip for {emp.first_name} {emp.last_name} — {run.year}/{run.month}\n")
            f.write(f"Net Salary: PKR {record.net_salary:,}\n")


# ─── Bank File ────────────────────────────────────────────────────────────────

async def generate_bank_file(
    tenant_id: str, run_id: str, db: AsyncSession
) -> bytes:
    """Generate IBFT Excel file for bulk salary disbursement."""
    run = await get_payroll_run_by_id(tenant_id, run_id, db)
    if run.status not in ("approved", "paid"):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "Bank file can only be generated for approved or paid payroll runs.",
        )

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status.HTTP_500_INTERNAL_SERVER_ERROR,
            "openpyxl is required for bank file generation.",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll_{run.year}_{run.month:02d}"

    headers = [
        "S.No", "Employee Code", "Employee Name",
        "Bank Name", "Account Title", "Account Number",
        "IBAN", "Net Salary (PKR)", "Reference",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E293B")

    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(hdr) + 4, 16)

    for idx, record in enumerate(run.records, 1):
        emp = record.employee
        bank = None
        if emp:
            bank_rows = await db.execute(
                select(BankDetails).where(
                    BankDetails.employee_id == emp.id,
                    BankDetails.is_primary  == True,
                    BankDetails.is_active   == True,
                )
            )
            bank = bank_rows.scalar_one_or_none()

        month_name = date(run.year, run.month, 1).strftime("%b%Y")
        reference  = f"SAL-{emp.employee_code if emp else 'N/A'}-{month_name}"

        row_data = [
            idx,
            emp.employee_code if emp else "",
            f"{emp.first_name} {emp.last_name}" if emp else "",
            bank.bank_name      if bank else "",
            bank.account_title  if bank else "",
            bank.account_number if bank else "",
            bank.iban           if bank else "",
            record.net_salary,
            reference,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            if col == 8:  # amount column
                cell.number_format = "#,##0"
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
