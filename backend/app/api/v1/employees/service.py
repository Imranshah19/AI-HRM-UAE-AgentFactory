"""
AI-HRMS — Employee service layer.

All business logic for employee CRUD. Routers delegate here; no SQL
in routers.

Design rules:
- Every query is scoped by tenant_id — never trust client-provided IDs alone.
- create_employee is fully atomic: Employee + User + Salary + Bank + Role
  are inserted in a single transaction; any failure rolls back everything.
- Audit log entries are written for every mutation.
- File paths are abstracted behind _storage_save() so switching from local
  to S3 only requires updating that one function.
"""

import hashlib
import mimetypes
import os
import re
import uuid
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any

import structlog
from fastapi import HTTPException, UploadFile, status
from sqlalchemy import func, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.v1.employees.schemas import (
    EmployeeCreateRequest,
    EmployeeDetail,
    EmployeeFilterParams,
    EmployeeListItem,
    EmployeeSalaryUpdateRequest,
    EmployeeStatusUpdateRequest,
    EmployeeUpdateRequest,
)
from app.core.config import settings
from app.core.security import hash_password
from app.models.audit import AuditLog
from app.models.compensation import BankDetails, SalaryStructure
from app.models.employee import Department, Designation, Employee, EmployeeDocument
from app.models.tenant import Role, Tenant, User, UserRole

logger = structlog.get_logger(__name__)

# ─── Constants ─────────────────────────────────────────────────────────────────

_ALLOWED_MIME_TYPES = {
    "application/pdf",
    "image/jpeg",
    "image/png",
    "image/jpg",
}
_MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


# ─── Helpers ───────────────────────────────────────────────────────────────────

def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _model_to_dict(obj: Any, exclude: set[str] | None = None) -> dict[str, Any]:
    """Shallow dict from an ORM object's __dict__ (excludes SQLAlchemy state)."""
    exclude = (exclude or set()) | {"_sa_instance_state"}
    return {k: v for k, v in obj.__dict__.items() if k not in exclude}


async def _write_audit_log(
    db:             AsyncSession,
    *,
    tenant_id:      uuid.UUID,
    user_id:        uuid.UUID | None,
    user_email:     str | None,
    action:         str,
    resource:       str,
    resource_id:    uuid.UUID | None = None,
    resource_label: str | None = None,
    old_values:     dict | None = None,
    new_values:     dict | None = None,
    changed_fields: list[str] | None = None,
) -> None:
    log = AuditLog(
        tenant_id=tenant_id,
        user_id=user_id,
        user_email=user_email,
        action=action,
        resource=resource,
        resource_id=resource_id,
        resource_label=resource_label,
        old_values=old_values,
        new_values=new_values,
        changed_fields=changed_fields,
    )
    db.add(log)


# ─── Storage abstraction ───────────────────────────────────────────────────────

async def _storage_save(
    file_bytes: bytes,
    relative_path: str,
) -> str:
    """
    Persist file to local disk (dev) or S3 (prod).
    Returns the public/accessible URL or relative path stored in the DB.
    """
    if settings.STORAGE_BACKEND == "s3":
        import boto3
        s3 = boto3.client(
            "s3",
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_REGION,
            endpoint_url=settings.AWS_S3_ENDPOINT_URL,
        )
        s3.put_object(
            Bucket=settings.AWS_S3_BUCKET,
            Key=relative_path,
            Body=file_bytes,
        )
        base = settings.AWS_S3_ENDPOINT_URL or f"https://{settings.AWS_S3_BUCKET}.s3.amazonaws.com"
        return f"{base}/{relative_path}"

    # Local storage
    full_path = os.path.join(settings.LOCAL_UPLOAD_DIR, relative_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, "wb") as fh:
        fh.write(file_bytes)
    return f"/uploads/{relative_path}"


async def _storage_delete(file_url: str) -> None:
    """Best-effort file deletion. Logs but does not raise on failure."""
    try:
        if settings.STORAGE_BACKEND == "s3":
            # derive key from URL
            import boto3
            s3 = boto3.client(
                "s3",
                aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
                region_name=settings.AWS_REGION,
                endpoint_url=settings.AWS_S3_ENDPOINT_URL,
            )
            key = file_url.split(f"{settings.AWS_S3_BUCKET}/", 1)[-1]
            s3.delete_object(Bucket=settings.AWS_S3_BUCKET, Key=key)
        else:
            rel = file_url.removeprefix("/uploads/")
            full = os.path.join(settings.LOCAL_UPLOAD_DIR, rel)
            if os.path.exists(full):
                os.remove(full)
    except Exception as exc:
        logger.warning("File deletion failed", url=file_url, error=str(exc))


# ─── Employee Code ─────────────────────────────────────────────────────────────

async def generate_employee_code(tenant_id: uuid.UUID, db: AsyncSession) -> str:
    """
    Return the next sequential employee code for the tenant: "EMP-0001".
    Uses COUNT of non-deleted employees to determine the next number.
    """
    result = await db.execute(
        select(func.count(Employee.id)).where(
            Employee.tenant_id == tenant_id,
            Employee.is_deleted.is_(False),
        )
    )
    count: int = result.scalar_one() or 0
    return f"EMP-{(count + 1):04d}"


# ─── Work Email ────────────────────────────────────────────────────────────────

async def generate_work_email(
    first_name: str,
    last_name:  str,
    tenant_slug: str,
    db:         AsyncSession,
    tenant_id:  uuid.UUID,
) -> str:
    """
    Derive a unique work email: john.doe@<slug>.ai-hrms.com
    If taken, append an incrementing suffix: john.doe2@…, john.doe3@…
    """
    base_local = f"{first_name.lower().strip()}.{last_name.lower().strip()}"
    # Strip non-alphanumeric except dots
    base_local = re.sub(r"[^a-z0-9.]", "", base_local)
    domain = f"{tenant_slug}.ai-hrms.com"

    # Find all existing emails matching this base pattern
    pattern = f"{base_local}%@{domain}"
    result = await db.execute(
        select(Employee.work_email).where(
            Employee.tenant_id == tenant_id,
            Employee.work_email.like(pattern),
            Employee.is_deleted.is_(False),
        )
    )
    existing = {row[0] for row in result.fetchall()}

    candidate = f"{base_local}@{domain}"
    if candidate not in existing:
        return candidate

    suffix = 2
    while True:
        candidate = f"{base_local}{suffix}@{domain}"
        if candidate not in existing:
            return candidate
        suffix += 1


# ─── Create Employee ──────────────────────────────────────────────────────────

async def create_employee(
    tenant_id:  uuid.UUID,
    tenant:     Tenant,
    data:       EmployeeCreateRequest,
    created_by: User,
    db:         AsyncSession,
) -> Employee:
    """
    Atomically create:
      1. User account (work_email as login)
      2. Employee record
      3. SalaryStructure (active)
      4. BankDetails (primary)
      5. UserRole assignment
    Fires welcome email Celery task after commit.
    """
    # ── 1. Validate uniqueness ────────────────────────────────────────────────
    if data.cnic:
        cnic_check = await db.execute(
            select(Employee.id).where(
                Employee.tenant_id == tenant_id,
                Employee.cnic == data.cnic,
                Employee.is_deleted.is_(False),
            )
        )
        if cnic_check.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"An employee with CNIC {data.cnic} already exists.",
            )

    # ── 2. Employee code ──────────────────────────────────────────────────────
    employee_code = data.employee_code or await generate_employee_code(tenant_id, db)

    # Check custom code uniqueness
    if data.employee_code:
        code_check = await db.execute(
            select(Employee.id).where(
                Employee.tenant_id == tenant_id,
                Employee.employee_code == employee_code,
                Employee.is_deleted.is_(False),
            )
        )
        if code_check.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Employee code '{employee_code}' is already in use.",
            )

    # ── 3. Work email ─────────────────────────────────────────────────────────
    work_email = await generate_work_email(
        data.first_name, data.last_name, tenant.slug, db, tenant_id
    )

    # ── 4. Temp password ──────────────────────────────────────────────────────
    temp_password = _generate_temp_password()

    # ── 5. User record ────────────────────────────────────────────────────────
    user = User(
        id=uuid.uuid4(),
        tenant_id=tenant_id,
        email=work_email,
        hashed_password=hash_password(temp_password),
        first_name=data.first_name,
        last_name=data.last_name,
        is_active=True,
        is_verified=False,
    )
    db.add(user)
    await db.flush()  # get user.id

    # ── 6. Employee record ────────────────────────────────────────────────────
    employee = Employee(
        id=uuid.uuid4(),
        tenant_id=tenant_id,
        user_id=user.id,
        employee_code=employee_code,
        first_name=data.first_name,
        last_name=data.last_name,
        father_name=data.father_name,
        cnic=data.cnic,
        personal_email=str(data.personal_email) if data.personal_email else None,
        work_email=work_email,
        phone=data.phone,
        gender=data.gender,
        dob=data.dob,
        marital_status=data.marital_status,
        nationality=data.nationality,
        address=data.address.model_dump() if data.address else None,
        emergency_contact=data.emergency_contact.model_dump() if data.emergency_contact else None,
        department_id=data.department_id,
        designation_id=data.designation_id,
        manager_id=data.manager_id,
        contract_type=data.contract_type,
        work_schedule=data.work_schedule,
        join_date=data.join_date,
        probation_end_date=data.probation_end_date,
        confirmation_date=data.confirmation_date,
        shift_id=data.shift_id,
        timezone=data.timezone,
        branch_location=data.branch_location,
        cost_center=data.cost_center,
        grade_level=data.grade_level,
        hr_notes=data.hr_notes,
        employment_status="active",
    )
    db.add(employee)
    await db.flush()  # get employee.id

    # ── 7. Salary structure ───────────────────────────────────────────────────
    salary = SalaryStructure(
        id=uuid.uuid4(),
        tenant_id=tenant_id,
        employee_id=employee.id,
        basic_salary=data.basic_salary,
        house_rent_allowance=data.house_rent_allowance,
        medical_allowance=data.medical_allowance,
        transport_allowance=data.transport_allowance,
        other_allowances=data.other_allowances,
        eobi_applicable=data.eobi_applicable,
        sessi_applicable=data.sessi_applicable,
        income_tax_applicable=data.income_tax_applicable,
        effective_from=data.join_date or date.today(),
        created_by=created_by.id,
    )
    db.add(salary)

    # ── 8. Bank details ───────────────────────────────────────────────────────
    if data.bank_name and data.account_number:
        bank = BankDetails(
            id=uuid.uuid4(),
            employee_id=employee.id,
            bank_name=data.bank_name,
            account_title=data.account_title or f"{data.first_name} {data.last_name}",
            account_number=data.account_number,
            iban=data.iban,
            is_primary=True,
        )
        db.add(bank)

    # ── 9. Role assignment ────────────────────────────────────────────────────
    if data.role_id:
        role_check = await db.execute(
            select(Role.id).where(
                Role.id == data.role_id,
                Role.tenant_id == tenant_id,
            )
        )
        if role_check.scalar_one_or_none():
            db.add(UserRole(user_id=user.id, role_id=data.role_id, assigned_by=created_by.id))

    # ── 10. Audit log ─────────────────────────────────────────────────────────
    await _write_audit_log(
        db,
        tenant_id=tenant_id,
        user_id=created_by.id,
        user_email=created_by.email,
        action="create",
        resource="employee",
        resource_id=employee.id,
        resource_label=f"{employee.full_name} ({employee_code})",
        new_values={
            "employee_code": employee_code,
            "work_email":    work_email,
            "contract_type": data.contract_type,
            "join_date":     str(data.join_date) if data.join_date else None,
        },
    )

    await db.flush()

    logger.info(
        "Employee created",
        employee_id=str(employee.id),
        employee_code=employee_code,
        tenant_id=str(tenant_id),
    )

    # ── 11. Queue welcome email (after commit, best-effort) ───────────────────
    # Import lazily to avoid circular imports at module load
    try:
        from app.tasks.employee_tasks import send_welcome_email
        send_welcome_email.delay(str(employee.id), temp_password)
    except Exception as exc:
        logger.warning("Could not queue welcome email", error=str(exc))

    return employee


def _generate_temp_password() -> str:
    """Generate a random 12-character temporary password."""
    import secrets
    import string
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    # Ensure at least one upper, lower, digit, special
    while True:
        pw = "".join(secrets.choice(alphabet) for _ in range(12))
        if (
            any(c.isupper() for c in pw)
            and any(c.islower() for c in pw)
            and any(c.isdigit() for c in pw)
            and any(c in "!@#$%" for c in pw)
        ):
            return pw


# ─── Query / List ──────────────────────────────────────────────────────────────

async def get_employees(
    tenant_id: uuid.UUID,
    filters:   EmployeeFilterParams,
    db:        AsyncSession,
) -> tuple[list[Employee], int]:
    """
    Return (employees, total_count) respecting all filter params.
    Employees are loaded with department and designation for the list view.
    """
    base = (
        select(Employee)
        .where(
            Employee.tenant_id == tenant_id,
            Employee.is_deleted.is_(False),
        )
        .options(
            selectinload(Employee.department),
            selectinload(Employee.designation),
        )
    )

    if filters.department_id:
        base = base.where(Employee.department_id == filters.department_id)
    if filters.designation_id:
        base = base.where(Employee.designation_id == filters.designation_id)
    if filters.manager_id:
        base = base.where(Employee.manager_id == filters.manager_id)
    if filters.status:
        base = base.where(Employee.employment_status == filters.status)
    if filters.contract_type:
        base = base.where(Employee.contract_type == filters.contract_type)

    if filters.search:
        term = filters.search.strip()
        ilike = f"%{term}%"
        base = base.where(
            or_(
                (Employee.first_name + " " + Employee.last_name).ilike(ilike),
                Employee.work_email.ilike(ilike),
                Employee.personal_email.ilike(ilike),
                Employee.cnic.ilike(ilike),
                Employee.employee_code.ilike(ilike),
            )
        )

    # Total count (before pagination)
    count_stmt = select(func.count()).select_from(base.subquery())
    total: int = (await db.execute(count_stmt)).scalar_one()

    # Paginated results
    offset = (filters.page - 1) * filters.page_size
    result = await db.execute(
        base.order_by(Employee.join_date.desc().nullslast(), Employee.created_at.desc())
            .offset(offset)
            .limit(filters.page_size)
    )
    employees = list(result.scalars().all())
    return employees, total


# ─── Get by ID ────────────────────────────────────────────────────────────────

async def get_employee_by_id(
    tenant_id:   uuid.UUID,
    employee_id: uuid.UUID,
    db:          AsyncSession,
) -> Employee:
    """
    Load a single employee with all relations needed for EmployeeDetail.
    Raises 404 if not found or belongs to a different tenant.
    """
    result = await db.execute(
        select(Employee)
        .where(
            Employee.id == employee_id,
            Employee.tenant_id == tenant_id,
            Employee.is_deleted.is_(False),
        )
        .options(
            selectinload(Employee.department),
            selectinload(Employee.designation),
            selectinload(Employee.manager),
            selectinload(Employee.documents),
            selectinload(Employee.salary_structures),
            selectinload(Employee.bank_details),
            selectinload(Employee.user).selectinload(User.user_roles).selectinload(
                UserRole.role
            ),
        )
    )
    employee = result.scalar_one_or_none()
    if employee is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Employee {employee_id} not found.",
        )
    return employee


# ─── Update Employee ───────────────────────────────────────────────────────────

async def update_employee(
    tenant_id:   uuid.UUID,
    employee_id: uuid.UUID,
    data:        EmployeeUpdateRequest,
    updated_by:  User,
    db:          AsyncSession,
) -> Employee:
    """Partial update of mutable employee fields."""
    employee = await get_employee_by_id(tenant_id, employee_id, db)

    old_values: dict[str, Any] = {}
    new_values: dict[str, Any] = {}
    changed_fields: list[str] = []

    for field, value in data.model_dump(exclude_unset=True).items():
        current = getattr(employee, field, None)
        # For JSONB fields (address, emergency_contact), compare serialised form
        if isinstance(value, dict) and isinstance(current, dict):
            if value != current:
                old_values[field] = current
                new_values[field] = value
                changed_fields.append(field)
                setattr(employee, field, value)
        elif value != current:
            old_values[field] = str(current) if current is not None else None
            new_values[field] = str(value) if value is not None else None
            changed_fields.append(field)
            setattr(employee, field, value)

    if changed_fields:
        await _write_audit_log(
            db,
            tenant_id=tenant_id,
            user_id=updated_by.id,
            user_email=updated_by.email,
            action="update",
            resource="employee",
            resource_id=employee.id,
            resource_label=f"{employee.full_name} ({employee.employee_code})",
            old_values=old_values,
            new_values=new_values,
            changed_fields=changed_fields,
        )
        db.add(employee)

    return employee


# ─── Update Status ────────────────────────────────────────────────────────────

async def update_employee_status(
    tenant_id:   uuid.UUID,
    employee_id: uuid.UUID,
    data:        EmployeeStatusUpdateRequest,
    updated_by:  User,
    db:          AsyncSession,
) -> Employee:
    employee = await get_employee_by_id(tenant_id, employee_id, db)

    old_status = employee.employment_status
    employee.employment_status = data.employment_status

    if data.employment_status in ("terminated", "resigned"):
        employee.termination_date   = data.effective_date or date.today()
        employee.termination_reason = data.reason

        # Deactivate the linked user account
        if employee.user_id:
            await db.execute(
                update(User)
                .where(User.id == employee.user_id)
                .values(is_active=False)
            )

        # Trigger offboarding workflow (Celery, best-effort)
        try:
            from app.tasks.employee_tasks import trigger_offboarding_workflow
            trigger_offboarding_workflow.delay(str(employee.id), data.reason or "")
        except Exception as exc:
            logger.warning("Could not queue offboarding task", error=str(exc))

    db.add(employee)

    await _write_audit_log(
        db,
        tenant_id=tenant_id,
        user_id=updated_by.id,
        user_email=updated_by.email,
        action="update",
        resource="employee",
        resource_id=employee.id,
        resource_label=f"{employee.full_name} ({employee.employee_code})",
        old_values={"employment_status": old_status},
        new_values={"employment_status": data.employment_status, "reason": data.reason},
        changed_fields=["employment_status"],
    )

    logger.info(
        "Employee status updated",
        employee_id=str(employee.id),
        old=old_status,
        new=data.employment_status,
    )
    return employee


# ─── Salary Update ────────────────────────────────────────────────────────────

async def update_employee_salary(
    tenant_id:   uuid.UUID,
    employee_id: uuid.UUID,
    data:        EmployeeSalaryUpdateRequest,
    updated_by:  User,
    db:          AsyncSession,
) -> SalaryStructure:
    """
    Close the current active salary structure and create a new one.
    This preserves full salary history.
    """
    # Verify employee exists and belongs to tenant
    employee = await get_employee_by_id(tenant_id, employee_id, db)

    # Close current active salary
    await db.execute(
        update(SalaryStructure)
        .where(
            SalaryStructure.employee_id == employee_id,
            SalaryStructure.effective_to.is_(None),
        )
        .values(effective_to=data.effective_from)
    )

    new_salary = SalaryStructure(
        id=uuid.uuid4(),
        tenant_id=tenant_id,
        employee_id=employee_id,
        basic_salary=data.basic_salary,
        house_rent_allowance=data.house_rent_allowance,
        medical_allowance=data.medical_allowance,
        transport_allowance=data.transport_allowance,
        other_allowances=data.other_allowances,
        eobi_applicable=data.eobi_applicable,
        sessi_applicable=data.sessi_applicable,
        income_tax_applicable=data.income_tax_applicable,
        effective_from=data.effective_from,
        effective_to=None,
        created_by=updated_by.id,
        revision_note=data.revision_note,
    )
    db.add(new_salary)

    await _write_audit_log(
        db,
        tenant_id=tenant_id,
        user_id=updated_by.id,
        user_email=updated_by.email,
        action="update",
        resource="salary_structure",
        resource_id=employee.id,
        resource_label=f"{employee.full_name} ({employee.employee_code})",
        new_values={
            "basic_salary":  data.basic_salary,
            "effective_from": str(data.effective_from),
        },
        changed_fields=["basic_salary", "effective_from"],
    )

    await db.flush()
    return new_salary


# ─── Document Upload ──────────────────────────────────────────────────────────

async def upload_employee_document(
    tenant_id:   uuid.UUID,
    employee_id: uuid.UUID,
    doc_type:    str,
    file:        UploadFile,
    uploaded_by: User,
    db:          AsyncSession,
) -> EmployeeDocument:
    """Validate, store, and record an employee document."""
    # Verify employee
    employee = await get_employee_by_id(tenant_id, employee_id, db)

    # Read file
    content = await file.read()

    # Size check
    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File exceeds the 5 MB limit. Uploaded size: {len(content):,} bytes.",
        )

    # MIME check
    mime_type, _ = mimetypes.guess_type(file.filename or "")
    if not mime_type:
        mime_type = file.content_type or "application/octet-stream"
    if mime_type not in _ALLOWED_MIME_TYPES:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail=f"File type '{mime_type}' is not allowed. Accepted: PDF, JPG, PNG.",
        )

    # Generate unique filename
    ext       = os.path.splitext(file.filename or "file")[1] or ".pdf"
    timestamp = int(_utcnow().timestamp())
    safe_name = f"{doc_type}_{timestamp}{ext}"
    rel_path  = f"{tenant_id}/{employee_id}/{safe_name}"

    file_url = await _storage_save(content, rel_path)

    doc = EmployeeDocument(
        id=uuid.uuid4(),
        employee_id=employee_id,
        doc_type=doc_type,
        doc_name=f"{doc_type.replace('_', ' ').title()} — {employee.full_name}",
        file_url=file_url,
        file_name=file.filename or safe_name,
        file_size_bytes=len(content),
        mime_type=mime_type,
    )
    db.add(doc)

    await _write_audit_log(
        db,
        tenant_id=tenant_id,
        user_id=uploaded_by.id,
        user_email=uploaded_by.email,
        action="create",
        resource="employee_document",
        resource_id=doc.id,
        resource_label=f"{doc.doc_name}",
        new_values={"doc_type": doc_type, "file_name": file.filename},
    )

    await db.flush()
    return doc


# ─── Document Delete ──────────────────────────────────────────────────────────

async def delete_employee_document(
    tenant_id:   uuid.UUID,
    employee_id: uuid.UUID,
    doc_id:      uuid.UUID,
    deleted_by:  User,
    db:          AsyncSession,
) -> None:
    """Soft-delete a document and best-effort remove the stored file."""
    # Verify employee belongs to tenant
    await get_employee_by_id(tenant_id, employee_id, db)

    result = await db.execute(
        select(EmployeeDocument).where(
            EmployeeDocument.id == doc_id,
            EmployeeDocument.employee_id == employee_id,
            EmployeeDocument.is_deleted.is_(False),
        )
    )
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Document {doc_id} not found.",
        )

    doc.is_deleted = True
    db.add(doc)

    await _write_audit_log(
        db,
        tenant_id=tenant_id,
        user_id=deleted_by.id,
        user_email=deleted_by.email,
        action="delete",
        resource="employee_document",
        resource_id=doc.id,
        resource_label=doc.doc_name,
    )

    await _storage_delete(doc.file_url)


# ─── Org Chart ─────────────────────────────────────────────────────────────────

async def get_org_chart(
    tenant_id: uuid.UUID,
    db:        AsyncSession,
    max_depth: int = 5,
) -> list[dict]:
    """
    Return a nested org-chart tree.
    Top-level nodes are employees with no manager (manager_id IS NULL).
    """
    result = await db.execute(
        select(Employee)
        .where(
            Employee.tenant_id == tenant_id,
            Employee.is_deleted.is_(False),
            Employee.employment_status == "active",
        )
        .options(
            selectinload(Employee.department),
            selectinload(Employee.designation),
        )
    )
    all_employees: list[Employee] = list(result.scalars().all())

    # Index by id for O(1) lookup
    emp_map: dict[uuid.UUID, dict] = {}
    for emp in all_employees:
        emp_map[emp.id] = {
            "id":            str(emp.id),
            "employee_code": emp.employee_code,
            "full_name":     emp.full_name,
            "designation":   emp.designation.name if emp.designation else None,
            "department":    emp.department.name if emp.department else None,
            "profile_photo_url": emp.profile_photo_url,
            "manager_id":    str(emp.manager_id) if emp.manager_id else None,
            "children":      [],
        }

    roots: list[dict] = []
    for emp in all_employees:
        node = emp_map[emp.id]
        if emp.manager_id and emp.manager_id in emp_map:
            emp_map[emp.manager_id]["children"].append(node)
        else:
            roots.append(node)

    return roots


# ─── Employee Excel Export ─────────────────────────────────────────────────────

async def export_employees_excel(
    tenant_id: uuid.UUID,
    filters:   EmployeeFilterParams,
    db:        AsyncSession,
) -> bytes:
    """
    Export filtered employees to an in-memory Excel workbook.
    Returns raw bytes; the router sets the response headers.
    """
    # Fetch ALL matching employees (no pagination for export)
    full_filters = filters.model_copy(update={"page": 1, "page_size": 10_000})
    employees, _ = await get_employees(tenant_id, full_filters, db)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export requires openpyxl. Install it with: pip install openpyxl",
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = [
        "Employee Code", "First Name", "Last Name", "Work Email",
        "Personal Email", "Phone", "CNIC", "Gender", "Date of Birth",
        "Department", "Designation", "Contract Type", "Status",
        "Join Date", "Probation End", "Branch / Location",
    ]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = 20

    for row_num, emp in enumerate(employees, start=2):
        ws.cell(row=row_num, column=1,  value=emp.employee_code)
        ws.cell(row=row_num, column=2,  value=emp.first_name)
        ws.cell(row=row_num, column=3,  value=emp.last_name)
        ws.cell(row=row_num, column=4,  value=emp.work_email)
        ws.cell(row=row_num, column=5,  value=emp.personal_email)
        ws.cell(row=row_num, column=6,  value=emp.phone)
        ws.cell(row=row_num, column=7,  value=emp.cnic)
        ws.cell(row=row_num, column=8,  value=emp.gender)
        ws.cell(row=row_num, column=9,  value=str(emp.dob) if emp.dob else None)
        ws.cell(row=row_num, column=10, value=emp.department.name if emp.department else None)
        ws.cell(row=row_num, column=11, value=emp.designation.name if emp.designation else None)
        ws.cell(row=row_num, column=12, value=emp.contract_type)
        ws.cell(row=row_num, column=13, value=emp.employment_status)
        ws.cell(row=row_num, column=14, value=str(emp.join_date) if emp.join_date else None)
        ws.cell(row=row_num, column=15, value=str(emp.probation_end_date) if emp.probation_end_date else None)
        ws.cell(row=row_num, column=16, value=emp.branch_location)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─── Active salary helper ─────────────────────────────────────────────────────

async def get_active_salary(
    employee_id: uuid.UUID,
    db:          AsyncSession,
) -> SalaryStructure | None:
    result = await db.execute(
        select(SalaryStructure).where(
            SalaryStructure.employee_id == employee_id,
            SalaryStructure.effective_to.is_(None),
        )
    )
    return result.scalar_one_or_none()
